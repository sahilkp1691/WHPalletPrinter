import io
import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Alignment, Font

from ..models import normalize_art_num

COL_CARTON = 1
COL_STOCK = 2
COL_TOTAL_QTY = 3
COL_QTY_PER_CARTON = 4
COL_NUM_CARTONS = 5

HEADERS = (
    "Carton Number",
    "Stock code",
    "Total Quantity",
    "Qty/Carton",
    "Number of Cartons",
)


@dataclass
class PacklistLine:
    row_num: int
    carton_spec: str
    carton_ids: list[int]
    stock_code: str
    total_qty: int
    qty_per_carton: int
    num_cartons: int
    warnings: list[str] = field(default_factory=list)


@dataclass
class PacklistParseResult:
    lines: list[PacklistLine]
    carton_index: dict[int, list[PacklistLine]]
    errors: list[str]
    warnings: list[str]
    stats: dict[str, int]


def parse_carton_scan(raw: str) -> int | None:
    text = (raw or "").strip()
    if not text:
        return None
    if not re.fullmatch(r"\d+", text):
        return None
    return int(text, 10)


def parse_carton_spec(spec: str) -> list[int] | None:
    text = (spec or "").strip()
    if not text:
        return None

    if "-" in text:
        parts = text.split("-", 1)
        if len(parts) != 2:
            return None
        start = parse_carton_scan(parts[0].strip())
        end = parse_carton_scan(parts[1].strip())
        if start is None or end is None or start > end:
            return None
        return list(range(start, end + 1))

    carton_id = parse_carton_scan(text)
    if carton_id is None:
        return None
    return [carton_id]


def _is_header_row(ws, row_num: int) -> bool:
    val = ws.cell(row_num, COL_CARTON).value
    if val is None:
        return False
    text = str(val).strip().lower()
    if not text:
        return False
    if parse_carton_spec(text) is not None:
        return False
    return True


def _merged_cell_map(ws) -> dict[tuple[int, int], object]:
    merged: dict[tuple[int, int], object] = {}
    for merge_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merge_range.bounds
        anchor = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged[(r, c)] = anchor
    return merged


def _cell_value(ws, merged: dict, row: int, col: int):
    val = ws.cell(row, col).value
    if val is None or (isinstance(val, str) and not val.strip()):
        return merged.get((row, col))
    return val


def _parse_int(value, field_name: str, row_num: int) -> tuple[int | None, str | None]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, f"Row {row_num}: {field_name} is empty"
    try:
        parsed = int(float(value))
    except (TypeError, ValueError):
        return None, f"Row {row_num}: invalid {field_name}"
    return parsed, None


def parse_packlist_xlsx(file_bytes: bytes) -> PacklistParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    merged = _merged_cell_map(ws)

    start_row = 1
    if ws.max_row >= 1 and _is_header_row(ws, 1):
        start_row = 2

    lines: list[PacklistLine] = []
    carton_index: dict[int, list[PacklistLine]] = {}
    errors: list[str] = []
    warnings: list[str] = []
    last_carton_spec: str | None = None

    for row_num in range(start_row, ws.max_row + 1):
        carton_raw = _cell_value(ws, merged, row_num, COL_CARTON)
        stock_raw = _cell_value(ws, merged, row_num, COL_STOCK)
        total_raw = _cell_value(ws, merged, row_num, COL_TOTAL_QTY)
        qty_ctn_raw = _cell_value(ws, merged, row_num, COL_QTY_PER_CARTON)
        num_ctn_raw = _cell_value(ws, merged, row_num, COL_NUM_CARTONS)

        if all(
            v is None or (isinstance(v, str) and not str(v).strip())
            for v in (carton_raw, stock_raw, total_raw, qty_ctn_raw, num_ctn_raw)
        ):
            continue

        carton_spec = str(carton_raw or "").strip()
        stock_code = normalize_art_num(str(stock_raw or ""))

        if not carton_spec and stock_code and last_carton_spec:
            carton_spec = last_carton_spec
        if carton_spec:
            last_carton_spec = carton_spec

        if not carton_spec:
            errors.append(f"Row {row_num}: Carton Number is empty")
            continue
        if not stock_code:
            errors.append(f"Row {row_num}: Stock code is empty")
            continue

        carton_ids = parse_carton_spec(carton_spec)
        if not carton_ids:
            errors.append(f"Row {row_num}: invalid Carton Number '{carton_spec}'")
            continue

        qty_per_carton, qty_err = _parse_int(qty_ctn_raw, "Qty/Carton", row_num)
        if qty_err:
            errors.append(qty_err)
            continue
        if qty_per_carton <= 0:
            errors.append(f"Row {row_num}: Qty/Carton must be > 0")
            continue

        total_qty, total_err = _parse_int(total_raw, "Total Quantity", row_num)
        if total_err:
            errors.append(total_err)
            continue

        num_cartons, num_err = _parse_int(num_ctn_raw, "Number of Cartons", row_num)
        if num_err:
            errors.append(num_err)
            continue
        if num_cartons <= 0:
            errors.append(f"Row {row_num}: Number of Cartons must be > 0")
            continue

        row_warnings: list[str] = []
        expected_from_range = len(carton_ids) * qty_per_carton
        expected_from_num = num_cartons * qty_per_carton
        if total_qty != expected_from_range:
            msg = (
                f"Row {row_num}: Total Quantity ({total_qty}) does not match "
                f"carton range count × Qty/Carton ({expected_from_range})"
            )
            row_warnings.append(msg)
            warnings.append(msg)
        if len(carton_ids) != num_cartons:
            msg = (
                f"Row {row_num}: carton range length ({len(carton_ids)}) "
                f"differs from Number of Cartons ({num_cartons})"
            )
            row_warnings.append(msg)
            warnings.append(msg)
        if total_qty != expected_from_num:
            msg = (
                f"Row {row_num}: Total Quantity ({total_qty}) does not match "
                f"Number of Cartons × Qty/Carton ({expected_from_num})"
            )
            if msg not in warnings:
                row_warnings.append(msg)
                warnings.append(msg)

        line = PacklistLine(
            row_num=row_num,
            carton_spec=carton_spec,
            carton_ids=carton_ids,
            stock_code=stock_code,
            total_qty=total_qty,
            qty_per_carton=qty_per_carton,
            num_cartons=num_cartons,
            warnings=row_warnings,
        )

        lines.append(line)
        for cid in carton_ids:
            carton_index.setdefault(cid, []).append(line)

    return PacklistParseResult(
        lines=lines,
        carton_index=carton_index,
        errors=errors,
        warnings=warnings,
        stats={
            "rows": len(lines),
            "cartons": len(carton_index),
        },
    )


def create_packlist_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packlist"
    ws.append(list(HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.append(["016-019", "PK1400", 48, 12, 4])
    ws.append(["020", "6PK1050", 6, 6, 1])
    ws.append([None, "PK2200", 6, 6, 1])
    ws.append(["022", "PK2200", 24, 24, 1])

    ws.merge_cells("A3:A4")
    ws.merge_cells("E3:E4")

    for col, width in zip("ABCDE", (16, 14, 16, 12, 18)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
