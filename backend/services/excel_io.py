import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session

from ..models import ArticleQtyCarton, normalize_art_num

ART_NUM_HEADER = "Art Num"
QTY_HEADER = "Qty/Carton"
REQUIRED_HEADERS = (ART_NUM_HEADER, QTY_HEADER)


def export_articles_xlsx(db: Session) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articles"
    ws.append(list(REQUIRED_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    rows = db.query(ArticleQtyCarton).order_by(ArticleQtyCarton.art_num).all()
    for row in rows:
        ws.append([row.art_num, row.qty_per_carton])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_articles_xlsx(db: Session, file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": ["Worksheet is empty"]}

    header_map = {_normalize_header(h): idx for idx, h in enumerate(header) if h is not None}
    missing = [h for h in REQUIRED_HEADERS if _normalize_header(h) not in header_map]
    if missing:
        return {
            "inserted": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [f"Missing required column(s): {', '.join(missing)}"],
        }

    art_idx = header_map[_normalize_header(ART_NUM_HEADER)]
    qty_idx = header_map[_normalize_header(QTY_HEADER)]

    inserted = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows_iter, start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        raw_art = row[art_idx] if art_idx < len(row) else None
        raw_qty = row[qty_idx] if qty_idx < len(row) else None

        art_num = normalize_art_num(str(raw_art or ""))
        if not art_num:
            skipped += 1
            errors.append(f"Row {row_num}: Art Num is empty")
            continue

        try:
            qty_per_carton = int(raw_qty)
        except (TypeError, ValueError):
            skipped += 1
            errors.append(f"Row {row_num}: invalid Qty/Carton for {art_num}")
            continue

        if qty_per_carton <= 0:
            skipped += 1
            errors.append(f"Row {row_num}: Qty/Carton must be > 0 for {art_num}")
            continue

        existing = db.get(ArticleQtyCarton, art_num)
        now = datetime.now(timezone.utc)
        if existing:
            existing.qty_per_carton = qty_per_carton
            existing.updated_at = now
            updated += 1
        else:
            db.add(
                ArticleQtyCarton(
                    art_num=art_num,
                    qty_per_carton=qty_per_carton,
                    updated_at=now,
                )
            )
            inserted += 1

    db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


def _normalize_header(value) -> str:
    return str(value or "").strip().lower()


def create_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articles"
    ws.append(list(REQUIRED_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["PK1400", 12])
    ws.append(["6PK1050", 6])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
